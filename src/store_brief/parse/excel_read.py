"""Load Excel workbooks with merged-cell expansion for cleaner table extraction."""
from __future__ import annotations

import zipfile
from pathlib import Path

import pandas as pd

from store_brief.parse.excel_tables import _cell_str, _drop_empty_columns
from store_brief.parse.layout_schema import RawSheet

MAX_SHEET_ROWS = 1200
MAX_SHEETS = 8


def _apply_merges(grid: list[list], ws) -> None:
    for merge_range in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merge_range.bounds
        value = ws.cell(min_row, min_col).value
        for r in range(min_row - 1, min_row - 1 + (max_row - min_row + 1)):
            if r >= len(grid):
                continue
            for c in range(min_col - 1, min_col - 1 + (max_col - min_col + 1)):
                if c >= len(grid[r]):
                    continue
                grid[r][c] = value


def _sheet_from_openpyxl(ws) -> pd.DataFrame:
    rows: list[list] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
        if len(rows) >= MAX_SHEET_ROWS:
            break
    if not rows:
        return pd.DataFrame()

    width = max(len(row) for row in rows)
    grid = [list(row) + [None] * (width - len(row)) for row in rows]
    _apply_merges(grid, ws)
    return pd.DataFrame(grid)


def read_workbook(path: Path) -> dict[str, pd.DataFrame]:
    lower = str(path).lower()
    if lower.endswith(".csv"):
        return {"_": pd.read_csv(path, header=None)}

    if lower.endswith((".xlsx", ".xlsm")) and zipfile.is_zipfile(path):
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=False)
        try:
            return {name: _sheet_from_openpyxl(wb[name]) for name in wb.sheetnames}
        finally:
            wb.close()

    if lower.endswith(".xlsb"):
        return pd.read_excel(path, sheet_name=None, engine="pyxlsb", header=None)
    if lower.endswith(".xls"):
        return pd.read_excel(path, sheet_name=None, engine="xlrd", header=None)
    return pd.read_excel(path, sheet_name=None, engine="openpyxl", header=None)


def dataframe_to_raw_rows(df: pd.DataFrame) -> list[list[str]]:
    if df.empty:
        return []
    rows = df.map(_cell_str).values.tolist()
    return _drop_empty_columns(rows)


def workbook_to_raw_sheets(sheets: dict[str, pd.DataFrame]) -> list[RawSheet]:
    out: list[RawSheet] = []
    for i, (name, df) in enumerate(sheets.items()):
        if i >= MAX_SHEETS:
            break
        rows = dataframe_to_raw_rows(df)
        if rows:
            out.append(RawSheet(sheet=str(name), rows=rows))
    return out


def read_raw_sheets(path: Path) -> list[RawSheet]:
    return workbook_to_raw_sheets(read_workbook(path))


def grid_snippet(rows: list[list[str]], *, max_rows: int = 15, max_cols: int = 12) -> str:
    lines: list[str] = []
    for i, row in enumerate(rows[:max_rows]):
        cells = [c for c in row[:max_cols] if c]
        if not cells:
            continue
        lines.append(f"R{i}: " + " | ".join(cells))
    return "\n".join(lines)
